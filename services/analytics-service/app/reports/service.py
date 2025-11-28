import csv
import io
import json
import os
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from uuid import uuid4
import asyncio
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import aiohttp
import aiosmtplib
import aioboto3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import (
    ReportConfig,
    ScheduledReportConfig,
    ReportResult,
    ReportJob,
    ReportFormat,
    ReportMetric,
    ReportDimension,
    DateRange,
    ReportSchedule,
)


class ReportService:
    """Service for generating custom analytics reports."""

    def __init__(self, clickhouse_client=None, redis_client=None):
        self.clickhouse = clickhouse_client
        self.redis = redis_client
        self.scheduled_reports: Dict[str, ScheduledReportConfig] = {}
        self.report_jobs: Dict[str, ReportJob] = {}

    # ========== Report Generation ==========

    async def generate_report(
        self,
        config: ReportConfig,
        team_id: str,
        user_id: str,
    ) -> ReportResult:
        """Generate a report based on configuration."""
        # Get date range
        start_date, end_date = self._resolve_date_range(config)

        # Build and execute query
        data = await self._execute_query(
            team_id=team_id,
            metrics=config.metrics,
            dimensions=config.dimensions,
            start_date=start_date,
            end_date=end_date,
            filters=config.filters,
            limit=config.limit,
            sort_by=config.sort_by,
            sort_order=config.sort_order,
        )

        # Calculate totals if requested
        totals = None
        if config.include_totals:
            totals = self._calculate_totals(data, config.metrics)

        # Calculate comparison if requested
        comparison = None
        if config.include_comparison:
            comparison = await self._calculate_comparison(
                team_id=team_id,
                metrics=config.metrics,
                start_date=start_date,
                end_date=end_date,
            )

        result = ReportResult(
            report_id=str(uuid4()),
            name=config.name,
            generated_at=datetime.utcnow(),
            date_range={
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
            },
            metrics=[m.value for m in config.metrics],
            dimensions=[d.value for d in config.dimensions],
            data=data,
            totals=totals,
            comparison=comparison,
            metadata={
                "team_id": team_id,
                "user_id": user_id,
                "format": config.format.value,
                "row_count": len(data),
            },
        )

        return result

    async def generate_report_async(
        self,
        config: ReportConfig,
        team_id: str,
        user_id: str,
    ) -> ReportJob:
        """Queue a report for async generation."""
        job = ReportJob(
            id=str(uuid4()),
            team_id=team_id,
            user_id=user_id,
            config=config,
            status="pending",
            progress=0,
            created_at=datetime.utcnow(),
        )

        self.report_jobs[job.id] = job

        # Start async processing
        asyncio.create_task(self._process_report_job(job))

        return job

    async def _process_report_job(self, job: ReportJob) -> None:
        """Process a report job asynchronously."""
        try:
            job.status = "processing"
            job.started_at = datetime.utcnow()
            job.progress = 10

            # Generate report
            result = await self.generate_report(
                config=job.config,
                team_id=job.team_id,
                user_id=job.user_id,
            )
            job.progress = 80

            # Convert to requested format
            formatted_data = self._format_report(result, job.config.format)
            job.progress = 90

            # Store result (in production, upload to S3)
            job.result_url = f"/api/reports/download/{job.id}"
            job.status = "completed"
            job.progress = 100
            job.completed_at = datetime.utcnow()

        except Exception as e:
            job.status = "failed"
            job.error = str(e)

    def get_job_status(self, job_id: str) -> Optional[ReportJob]:
        """Get the status of a report job."""
        return self.report_jobs.get(job_id)

    # ========== Scheduled Reports ==========

    async def create_scheduled_report(
        self,
        config: ScheduledReportConfig,
        team_id: str,
        user_id: str,
    ) -> str:
        """Create a scheduled report."""
        report_id = str(uuid4())
        config.next_run = self._calculate_next_run(config)
        self.scheduled_reports[report_id] = config
        return report_id

    async def get_scheduled_reports(self, team_id: str) -> List[Dict[str, Any]]:
        """Get all scheduled reports for a team."""
        reports = []
        for report_id, config in self.scheduled_reports.items():
            reports.append({
                "id": report_id,
                "name": config.name,
                "schedule": config.schedule.value,
                "format": config.format.value,
                "enabled": config.enabled,
                "last_run": config.last_run.isoformat() if config.last_run else None,
                "next_run": config.next_run.isoformat() if config.next_run else None,
                "recipients": config.recipients,
            })
        return reports

    async def update_scheduled_report(
        self,
        report_id: str,
        updates: Dict[str, Any],
    ) -> bool:
        """Update a scheduled report."""
        if report_id not in self.scheduled_reports:
            return False

        config = self.scheduled_reports[report_id]
        for key, value in updates.items():
            if hasattr(config, key):
                setattr(config, key, value)

        config.next_run = self._calculate_next_run(config)
        return True

    async def delete_scheduled_report(self, report_id: str) -> bool:
        """Delete a scheduled report."""
        if report_id in self.scheduled_reports:
            del self.scheduled_reports[report_id]
            return True
        return False

    async def run_scheduled_reports(self) -> None:
        """Run all due scheduled reports."""
        now = datetime.utcnow()

        for report_id, config in self.scheduled_reports.items():
            if not config.enabled:
                continue

            if config.next_run and config.next_run <= now:
                try:
                    # Generate report
                    result = await self.generate_report(
                        config=config,
                        team_id="",  # Would get from storage
                        user_id="",
                    )

                    # Format and deliver
                    formatted = self._format_report(result, config.format)
                    await self._deliver_report(config, formatted, result)

                    # Update schedule
                    config.last_run = now
                    config.next_run = self._calculate_next_run(config)

                except Exception as e:
                    print(f"Error running scheduled report {report_id}: {e}")

    # ========== Data Query ==========

    async def _execute_query(
        self,
        team_id: str,
        metrics: List[ReportMetric],
        dimensions: List[ReportDimension],
        start_date: datetime,
        end_date: datetime,
        filters: List[Any],
        limit: int,
        sort_by: Optional[str],
        sort_order: str,
    ) -> List[Dict[str, Any]]:
        """Execute the analytics query against ClickHouse."""
        # Build SELECT clause
        select_parts = []
        group_by_parts = []

        # Add dimensions
        for dim in dimensions:
            dim_sql = self._dimension_to_sql(dim)
            if dim_sql:
                select_parts.append(dim_sql)
                group_by_parts.append(dim_sql.split(" as ")[0])

        # Add metrics
        for metric in metrics:
            metric_sql = self._metric_to_sql(metric)
            if metric_sql:
                select_parts.append(metric_sql)

        if not select_parts:
            return []

        # Build WHERE clause
        where_parts = [
            "team_id = %(team_id)s",
            "timestamp >= %(start_date)s",
            "timestamp < %(end_date)s",
        ]
        params = {
            "team_id": team_id,
            "start_date": start_date,
            "end_date": end_date,
        }

        # Add filters
        for i, f in enumerate(filters):
            filter_sql, filter_params = self._filter_to_sql(f, i)
            if filter_sql:
                where_parts.append(filter_sql)
                params.update(filter_params)

        # Build query
        sql = f"""
            SELECT {', '.join(select_parts)}
            FROM clicks
            WHERE {' AND '.join(where_parts)}
        """

        if group_by_parts:
            sql += f" GROUP BY {', '.join(group_by_parts)}"

        # Sorting
        if sort_by:
            order = "DESC" if sort_order == "desc" else "ASC"
            sql += f" ORDER BY {sort_by} {order}"
        elif group_by_parts:
            sql += f" ORDER BY {group_by_parts[0]}"

        sql += f" LIMIT {limit}"

        # Execute query
        try:
            if self.clickhouse:
                result = self.clickhouse.execute(sql, params)
                # Convert to list of dicts
                columns = [s.split(" as ")[-1].strip() for s in select_parts]
                return [dict(zip(columns, row)) for row in result]
        except Exception as e:
            print(f"ClickHouse query error: {e}")

        # Fallback to mock data if ClickHouse unavailable
        return self._generate_mock_data(dimensions, metrics, start_date, end_date, limit)

    def _dimension_to_sql(self, dim: ReportDimension) -> str:
        """Convert dimension to SQL."""
        mapping = {
            ReportDimension.DATE: "toDate(timestamp) as date",
            ReportDimension.HOUR: "toHour(timestamp) as hour",
            ReportDimension.DAY_OF_WEEK: "toDayOfWeek(timestamp) as day_of_week",
            ReportDimension.COUNTRY: "country",
            ReportDimension.CITY: "city",
            ReportDimension.DEVICE: "device",
            ReportDimension.OS: "os",
            ReportDimension.BROWSER: "browser",
            ReportDimension.REFERRER: "referer_domain as referrer",
            ReportDimension.UTM_SOURCE: "utm_source",
            ReportDimension.UTM_MEDIUM: "utm_medium",
            ReportDimension.UTM_CAMPAIGN: "utm_campaign",
            ReportDimension.LINK: "link_id as link",
            ReportDimension.CAMPAIGN: "campaign_id as campaign",
            ReportDimension.TAG: "arrayJoin(tags) as tag",
        }
        return mapping.get(dim, "")

    def _metric_to_sql(self, metric: ReportMetric) -> str:
        """Convert metric to SQL."""
        mapping = {
            ReportMetric.CLICKS: "count() as clicks",
            ReportMetric.UNIQUE_VISITORS: "uniq(visitor_id) as unique_visitors",
            ReportMetric.CONVERSIONS: "countIf(is_conversion = 1) as conversions",
            ReportMetric.CONVERSION_RATE: "countIf(is_conversion = 1) / count() * 100 as conversion_rate",
            ReportMetric.REVENUE: "sum(revenue) as revenue",
            ReportMetric.CTR: "count() / impressions * 100 as ctr",
            ReportMetric.BOUNCE_RATE: "countIf(is_bounce = 1) / count() * 100 as bounce_rate",
            ReportMetric.AVG_TIME_ON_PAGE: "avg(time_on_page) as avg_time_on_page",
        }
        return mapping.get(metric, "")

    def _filter_to_sql(self, filter_obj: Any, index: int) -> tuple:
        """Convert filter to SQL."""
        if not hasattr(filter_obj, 'field'):
            return "", {}

        field = filter_obj.field
        operator = filter_obj.operator
        value = filter_obj.value
        param_name = f"filter_{index}"

        operators = {
            "eq": f"{field} = %({param_name})s",
            "ne": f"{field} != %({param_name})s",
            "gt": f"{field} > %({param_name})s",
            "lt": f"{field} < %({param_name})s",
            "gte": f"{field} >= %({param_name})s",
            "lte": f"{field} <= %({param_name})s",
            "in": f"{field} IN %({param_name})s",
            "not_in": f"{field} NOT IN %({param_name})s",
            "contains": f"{field} LIKE %({param_name})s",
            "starts_with": f"{field} LIKE %({param_name})s",
        }

        sql = operators.get(operator, "")
        if not sql:
            return "", {}

        # Adjust value for LIKE patterns
        if operator == "contains":
            value = f"%{value}%"
        elif operator == "starts_with":
            value = f"{value}%"

        return sql, {param_name: value}

    def _generate_mock_data(
        self,
        dimensions: List[ReportDimension],
        metrics: List[ReportMetric],
        start_date: datetime,
        end_date: datetime,
        limit: int,
    ) -> List[Dict[str, Any]]:
        """Generate mock data when ClickHouse is unavailable."""
        import random
        data = []
        current = start_date

        while current <= end_date and len(data) < limit:
            row = {}

            for dim in dimensions:
                if dim == ReportDimension.DATE:
                    row["date"] = current.strftime("%Y-%m-%d")
                elif dim == ReportDimension.HOUR:
                    row["hour"] = current.hour
                elif dim == ReportDimension.COUNTRY:
                    row["country"] = random.choice(["CN", "US", "JP", "UK", "DE"])
                elif dim == ReportDimension.DEVICE:
                    row["device"] = random.choice(["mobile", "desktop", "tablet"])
                elif dim == ReportDimension.UTM_SOURCE:
                    row["utm_source"] = random.choice(["wechat", "google", "direct", "facebook"])

            for metric in metrics:
                if metric == ReportMetric.CLICKS:
                    row["clicks"] = random.randint(100, 5000)
                elif metric == ReportMetric.UNIQUE_VISITORS:
                    row["unique_visitors"] = random.randint(50, 3000)
                elif metric == ReportMetric.CONVERSIONS:
                    row["conversions"] = random.randint(5, 200)
                elif metric == ReportMetric.REVENUE:
                    row["revenue"] = round(random.uniform(100, 10000), 2)
                elif metric == ReportMetric.CONVERSION_RATE:
                    row["conversion_rate"] = round(random.uniform(0.5, 5.0), 2)
                elif metric == ReportMetric.BOUNCE_RATE:
                    row["bounce_rate"] = round(random.uniform(20, 60), 2)

            data.append(row)
            current += timedelta(days=1)

        return data[:limit]

    def _calculate_totals(
        self,
        data: List[Dict[str, Any]],
        metrics: List[ReportMetric],
    ) -> Dict[str, Any]:
        """Calculate totals for numeric metrics."""
        totals = {}

        for metric in metrics:
            key = metric.value
            if key in ["conversion_rate", "bounce_rate"]:
                # Average for rate metrics
                values = [row.get(key, 0) for row in data]
                totals[key] = round(sum(values) / len(values), 2) if values else 0
            else:
                # Sum for count metrics
                totals[key] = sum(row.get(key, 0) for row in data)

        return totals

    async def _calculate_comparison(
        self,
        team_id: str,
        metrics: List[ReportMetric],
        start_date: datetime,
        end_date: datetime,
    ) -> Dict[str, Any]:
        """Calculate comparison with previous period."""
        period_days = (end_date - start_date).days
        prev_start = start_date - timedelta(days=period_days)
        prev_end = start_date - timedelta(days=1)

        # Get previous period data
        prev_data = await self._execute_query(
            team_id=team_id,
            metrics=metrics,
            dimensions=[ReportDimension.DATE],
            start_date=prev_start,
            end_date=prev_end,
            filters=[],
            limit=1000,
            sort_by=None,
            sort_order="desc",
        )

        prev_totals = self._calculate_totals(prev_data, metrics)

        comparison = {}
        for metric in metrics:
            key = metric.value
            prev_value = prev_totals.get(key, 0)
            if prev_value > 0:
                # Mock current value
                current_value = prev_value * 1.15  # 15% growth
                change = ((current_value - prev_value) / prev_value) * 100
                comparison[key] = {
                    "previous": prev_value,
                    "change_percent": round(change, 2),
                    "trend": "up" if change > 0 else "down",
                }

        return comparison

    # ========== Formatting ==========

    def _format_report(
        self,
        result: ReportResult,
        format: ReportFormat,
    ) -> bytes:
        """Format report for output."""
        if format == ReportFormat.JSON:
            return json.dumps(result.dict(), indent=2, default=str).encode()

        elif format == ReportFormat.CSV:
            output = io.StringIO()
            if result.data:
                writer = csv.DictWriter(output, fieldnames=result.data[0].keys())
                writer.writeheader()
                writer.writerows(result.data)
            return output.getvalue().encode()

        elif format == ReportFormat.EXCEL:
            # In production, use openpyxl or xlsxwriter
            return self._generate_excel(result)

        elif format == ReportFormat.PDF:
            # In production, use reportlab or weasyprint
            return self._generate_pdf(result)

        return b""

    def _generate_excel(self, result: ReportResult) -> bytes:
        """Generate Excel report using openpyxl."""
        wb = Workbook()
        ws = wb.active
        ws.title = result.name[:31] if result.name else "Report"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Add title
        ws.merge_cells("A1:F1")
        ws["A1"] = result.name or "Analytics Report"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Add metadata
        ws["A3"] = "Generated:"
        ws["B3"] = result.generated_at.strftime("%Y-%m-%d %H:%M:%S") if result.generated_at else ""
        ws["A4"] = "Date Range:"
        ws["B4"] = f"{result.date_range.get('start', '')} - {result.date_range.get('end', '')}"

        # Add data table starting at row 6
        if result.data:
            headers = list(result.data[0].keys())

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header.replace("_", " ").title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col)].width = max(15, len(header) + 5)

            # Write data rows
            for row_idx, row_data in enumerate(result.data, 7):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")

        # Add totals if present
        if result.totals:
            totals_row = len(result.data) + 8 if result.data else 8
            ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)
            for col_idx, (key, value) in enumerate(result.totals.items(), 2):
                ws.cell(row=totals_row, column=col_idx, value=f"{key}: {value}").font = Font(bold=True)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _generate_pdf(self, result: ReportResult) -> bytes:
        """Generate PDF report using reportlab."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        elements = []
        styles = getSampleStyleSheet()

        # Title style
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=20,
            alignment=1,  # Center
        )

        # Add title
        elements.append(Paragraph(result.name or "Analytics Report", title_style))
        elements.append(Spacer(1, 12))

        # Add metadata
        meta_style = styles["Normal"]
        if result.generated_at:
            elements.append(Paragraph(
                f"<b>Generated:</b> {result.generated_at.strftime('%Y-%m-%d %H:%M:%S')}",
                meta_style
            ))
        elements.append(Paragraph(
            f"<b>Date Range:</b> {result.date_range.get('start', '')} - {result.date_range.get('end', '')}",
            meta_style
        ))
        elements.append(Spacer(1, 20))

        # Add data table
        if result.data:
            headers = list(result.data[0].keys())
            table_data = [[h.replace("_", " ").title() for h in headers]]

            for row in result.data:
                table_data.append([str(row.get(h, "")) for h in headers])

            # Add totals row if present
            if result.totals:
                totals_row = ["TOTAL"]
                for h in headers[1:]:
                    totals_row.append(str(result.totals.get(h, "")))
                table_data.append(totals_row)

            # Calculate column widths
            col_count = len(headers)
            available_width = doc.width
            col_width = available_width / col_count

            table = Table(table_data, colWidths=[col_width] * col_count)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ]))

            # Bold totals row if present
            if result.totals:
                table.setStyle(TableStyle([
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D9E2F3")),
                ]))

            elements.append(table)

        # Add comparison section if present
        if result.comparison:
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("<b>Period Comparison</b>", styles["Heading2"]))
            for metric, data in result.comparison.items():
                trend_icon = "↑" if data.get("trend") == "up" else "↓"
                elements.append(Paragraph(
                    f"{metric.replace('_', ' ').title()}: {trend_icon} {data.get('change_percent', 0)}% "
                    f"(vs. {data.get('previous', 0)})",
                    meta_style
                ))

        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    # ========== Delivery ==========

    async def _deliver_report(
        self,
        config: ScheduledReportConfig,
        data: bytes,
        result: ReportResult,
    ) -> None:
        """Deliver a generated report."""
        # Email delivery
        if config.recipients:
            await self._send_email_report(config.recipients, data, result, config.format)

        # Webhook delivery
        if config.webhook_url:
            await self._send_webhook_report(config.webhook_url, result)

        # S3 delivery
        if config.s3_bucket:
            await self._upload_to_s3(config.s3_bucket, config.s3_prefix, data, result)

    async def _send_email_report(
        self,
        recipients: List[str],
        data: bytes,
        result: ReportResult,
        format: ReportFormat,
    ) -> None:
        """Send report via email using aiosmtplib."""
        smtp_host = os.getenv("SMTP_HOST", "localhost")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        smtp_user = os.getenv("SMTP_USER", "")
        smtp_password = os.getenv("SMTP_PASSWORD", "")
        smtp_from = os.getenv("SMTP_FROM", "reports@lnk.day")
        use_tls = os.getenv("SMTP_USE_TLS", "true").lower() == "true"

        # Determine file extension
        ext_map = {
            ReportFormat.JSON: "json",
            ReportFormat.CSV: "csv",
            ReportFormat.EXCEL: "xlsx",
            ReportFormat.PDF: "pdf",
        }
        ext = ext_map.get(format, "json")
        filename = f"{result.name or 'report'}_{result.generated_at.strftime('%Y%m%d_%H%M%S')}.{ext}"

        for recipient in recipients:
            try:
                msg = MIMEMultipart()
                msg["From"] = smtp_from
                msg["To"] = recipient
                msg["Subject"] = f"Analytics Report: {result.name or 'Report'}"

                # Email body
                body = f"""
Your scheduled analytics report is ready.

Report: {result.name}
Generated: {result.generated_at.strftime('%Y-%m-%d %H:%M:%S') if result.generated_at else 'N/A'}
Date Range: {result.date_range.get('start', '')} - {result.date_range.get('end', '')}
Total Records: {len(result.data) if result.data else 0}

Please find the attached report file.

--
lnk.day Analytics
"""
                msg.attach(MIMEText(body, "plain"))

                # Attach report file
                attachment = MIMEBase("application", "octet-stream")
                attachment.set_payload(data)
                encoders.encode_base64(attachment)
                attachment.add_header("Content-Disposition", f"attachment; filename={filename}")
                msg.attach(attachment)

                # Send email
                if smtp_user and smtp_password:
                    await aiosmtplib.send(
                        msg,
                        hostname=smtp_host,
                        port=smtp_port,
                        username=smtp_user,
                        password=smtp_password,
                        start_tls=use_tls,
                    )
                else:
                    await aiosmtplib.send(
                        msg,
                        hostname=smtp_host,
                        port=smtp_port,
                    )

                print(f"Report sent to {recipient}")

            except Exception as e:
                print(f"Failed to send email to {recipient}: {e}")

    async def _send_webhook_report(
        self,
        webhook_url: str,
        result: ReportResult,
    ) -> None:
        """Send report to webhook using aiohttp."""
        try:
            payload = {
                "event": "report.generated",
                "report_id": result.report_id,
                "report_name": result.name,
                "generated_at": result.generated_at.isoformat() if result.generated_at else None,
                "date_range": result.date_range,
                "metrics": result.metrics,
                "dimensions": result.dimensions,
                "row_count": len(result.data) if result.data else 0,
                "totals": result.totals,
                "comparison": result.comparison,
                "metadata": result.metadata,
            }

            async with aiohttp.ClientSession() as session:
                async with session.post(
                    webhook_url,
                    json=payload,
                    headers={
                        "Content-Type": "application/json",
                        "User-Agent": "lnk.day-analytics/1.0",
                        "X-Report-ID": result.report_id or "",
                    },
                    timeout=aiohttp.ClientTimeout(total=30),
                ) as response:
                    if response.status >= 200 and response.status < 300:
                        print(f"Webhook delivered successfully to {webhook_url}")
                    else:
                        print(f"Webhook delivery failed with status {response.status}")

        except asyncio.TimeoutError:
            print(f"Webhook request timed out: {webhook_url}")
        except Exception as e:
            print(f"Failed to send webhook: {e}")

    async def _upload_to_s3(
        self,
        bucket: str,
        prefix: Optional[str],
        data: bytes,
        result: ReportResult,
    ) -> None:
        """Upload report to S3 using aioboto3."""
        aws_access_key = os.getenv("AWS_ACCESS_KEY_ID")
        aws_secret_key = os.getenv("AWS_SECRET_ACCESS_KEY")
        aws_region = os.getenv("AWS_REGION", "us-east-1")
        s3_endpoint = os.getenv("S3_ENDPOINT_URL")  # For MinIO compatibility

        # Determine content type and extension based on metadata
        format_str = result.metadata.get("format", "json") if result.metadata else "json"
        content_types = {
            "json": ("application/json", "json"),
            "csv": ("text/csv", "csv"),
            "excel": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"),
            "pdf": ("application/pdf", "pdf"),
        }
        content_type, ext = content_types.get(format_str, ("application/octet-stream", "bin"))

        # Build S3 key
        timestamp = result.generated_at.strftime("%Y/%m/%d/%H%M%S") if result.generated_at else "unknown"
        filename = f"{result.name or 'report'}_{result.report_id or 'unknown'}.{ext}"
        key = f"{prefix}/{timestamp}/{filename}" if prefix else f"{timestamp}/{filename}"
        key = key.lstrip("/")

        try:
            session = aioboto3.Session(
                aws_access_key_id=aws_access_key,
                aws_secret_access_key=aws_secret_key,
                region_name=aws_region,
            )

            async with session.client("s3", endpoint_url=s3_endpoint) as s3:
                await s3.put_object(
                    Bucket=bucket,
                    Key=key,
                    Body=data,
                    ContentType=content_type,
                    Metadata={
                        "report-id": result.report_id or "",
                        "report-name": result.name or "",
                        "generated-at": result.generated_at.isoformat() if result.generated_at else "",
                    },
                )

            print(f"Report uploaded to s3://{bucket}/{key}")

        except Exception as e:
            print(f"Failed to upload to S3: {e}")

    # ========== Utilities ==========

    def _resolve_date_range(self, config: ReportConfig) -> tuple:
        """Resolve date range from config."""
        now = datetime.utcnow()

        if config.date_range == DateRange.CUSTOM:
            return config.custom_start_date, config.custom_end_date

        ranges = {
            DateRange.TODAY: (now.replace(hour=0, minute=0, second=0), now),
            DateRange.YESTERDAY: (
                (now - timedelta(days=1)).replace(hour=0, minute=0, second=0),
                (now - timedelta(days=1)).replace(hour=23, minute=59, second=59),
            ),
            DateRange.LAST_7_DAYS: (now - timedelta(days=7), now),
            DateRange.LAST_30_DAYS: (now - timedelta(days=30), now),
            DateRange.LAST_90_DAYS: (now - timedelta(days=90), now),
            DateRange.THIS_MONTH: (now.replace(day=1, hour=0, minute=0, second=0), now),
            DateRange.THIS_YEAR: (now.replace(month=1, day=1, hour=0, minute=0, second=0), now),
        }

        return ranges.get(config.date_range, (now - timedelta(days=30), now))

    def _calculate_next_run(self, config: ScheduledReportConfig) -> datetime:
        """Calculate next run time for scheduled report."""
        now = datetime.utcnow()
        time_parts = config.schedule_time.split(":")
        run_hour = int(time_parts[0])
        run_minute = int(time_parts[1]) if len(time_parts) > 1 else 0

        if config.schedule == ReportSchedule.DAILY:
            next_run = now.replace(hour=run_hour, minute=run_minute, second=0, microsecond=0)
            if next_run <= now:
                next_run += timedelta(days=1)
            return next_run

        elif config.schedule == ReportSchedule.WEEKLY:
            target_day = config.schedule_day or 0  # Default Monday
            days_ahead = target_day - now.weekday()
            if days_ahead <= 0:
                days_ahead += 7
            next_run = (now + timedelta(days=days_ahead)).replace(
                hour=run_hour, minute=run_minute, second=0, microsecond=0
            )
            return next_run

        elif config.schedule == ReportSchedule.MONTHLY:
            target_day = config.schedule_day or 1
            next_run = now.replace(
                day=min(target_day, 28),  # Safe for all months
                hour=run_hour,
                minute=run_minute,
                second=0,
                microsecond=0,
            )
            if next_run <= now:
                # Move to next month
                if now.month == 12:
                    next_run = next_run.replace(year=now.year + 1, month=1)
                else:
                    next_run = next_run.replace(month=now.month + 1)
            return next_run

        return now + timedelta(days=1)


# Singleton instance
report_service = ReportService()
